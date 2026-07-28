#!/usr/bin/env python3
"""
Generador/verificador de cargos — Paseo Nordelta (Ctas Ctes).

La fuente de verdad de cada cargo es la PESTAÑA DEL LOCAL (ahí las fórmulas
calculan alquiler indexado por IPC y expensas desde Expensas Predio). CARGOS
es la tabla que consume CUENTA CORRIENTE. Este script cruza las dos:

  - cargo en la pestaña del local pero no en CARGOS  → FALTA (se puede escribir)
  - cargo en CARGOS con total 0 y valor real en la pestaña → EN CERO (se corrige)
  - concepto sin calcular todavía en la pestaña (ej. expensas del mes corriente)
    → PENDIENTE (se informa, no se inventa)

    .venv/bin/python generar_cargos.py <ctas_ctes.xlsx> <AAAA-MM> [--escribir]

Sin --escribir SOLO reporta. Con --escribir corrige los EN CERO y agrega los
FALTA vía API (nota "generar_cargos.py"), nunca pisa un cargo con valor.
"""

import datetime
import re
import sys

import openpyxl

FILE_ID = "10BDmKvv2wY2M4lVYYab3NiNT04WLh5JjO-EhWI4tnNs"

# Pestaña del local → nombre del local en CARGOS/CUENTA CORRIENTE
TABS = {
    "Fabric": "Fabric", "Bigg": "Bigg", "Boss": "Boss",
    "Volta + Open 25": "Volta + Open", "Peak One": "Peak One",
    "Salon Multiespacios": "Salón (Alto)",
}
MESES = {1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
         7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC"}
# En las pestañas el detalle aparece con variantes; se normaliza a los conceptos
# que usa CARGOS. El IVA viene como línea aparte "IVA <concepto>".
CONCEPTOS = {
    "alquiler": "Alquiler",
    "diferencia alquiler (sin iva)": "Alquiler",   # mitad en efectivo de Bigg
    "diferencia alquiler": "Alquiler",
    "recupero de gastos": "Recupero de gastos",
    "rec gs fc": "Recupero de gastos",
    "servicios comunes": "Servicios comunes",
}


def etiqueta_mes(anio, num):
    return f"{MESES[num]}'{anio % 100}"


def cargos_de_pestania(wb, tab, anio, num):
    """Lee los cargos del mes en la pestaña del local: {concepto: (monto, iva)}."""
    ws = wb[tab]
    # Ubicar el bloque de cargos: fila de header con "Detalle" y columna "Egreso"
    col_detalle = col_egreso = header_row = None
    for r in range(1, 12):
        for c in range(1, 12):
            v = str(ws.cell(row=r, column=c).value or "").strip().lower()
            if v == "detalle":
                header_row, col_detalle = r, c
            elif v == "egreso" and header_row == r:
                col_egreso = c
        if col_detalle and col_egreso:
            break
    if not (col_detalle and col_egreso):
        return None  # pestaña sin bloque reconocible (plantilla vacía)

    objetivo = etiqueta_mes(anio, num)
    out, conteo = {}, {}
    for r in range(header_row + 1, ws.max_row + 1):
        mes = str(ws.cell(row=r, column=1).value or "").upper().replace(" ", "")
        if mes != objetivo.replace(" ", ""):
            continue
        detalle = str(ws.cell(row=r, column=col_detalle).value or "").strip()
        monto = ws.cell(row=r, column=col_egreso).value
        monto = float(monto) if isinstance(monto, (int, float)) else 0.0
        d = detalle.lower()
        # Un mismo concepto puede tener VARIAS filas en el mes (Bigg: alquiler
        # facturado + mitad en efectivo + diferencias recuperadas): se SUMAN.
        if d.startswith("iva "):
            base = CONCEPTOS.get(d[4:].strip())
            if base and base in out:
                out[base] = (out[base][0], out[base][1] + monto)
        else:
            base = CONCEPTOS.get(d)
            if base:
                monto_prev, iva_prev = out.get(base, (0.0, 0.0))
                out[base] = (monto_prev + monto, iva_prev)
                conteo[base] = conteo.get(base, 0) + 1
    return out, conteo


def cargos_existentes(wb, anio, num):
    """CARGOS del mes: {(local, concepto): (fila_sheet, total)}."""
    out = {}
    for i, r in enumerate(wb["CARGOS"].iter_rows(min_row=1, values_only=True), start=1):
        if isinstance(r[0], datetime.datetime) and (r[0].year, r[0].month) == (anio, num):
            out[(str(r[1]).strip(), str(r[2]).strip())] = (i, float(r[5] or 0))
    return out


def main():
    argv = sys.argv[1:]
    escribir = "--escribir" in argv
    if escribir:
        argv.remove("--escribir")
    if len(argv) != 2:
        sys.exit(__doc__)
    xlsx, mes = argv
    m = re.fullmatch(r"(\d{4})-(\d{2})", mes)
    if not m or not 1 <= int(m.group(2)) <= 12:
        sys.exit(f"El mes va como AAAA-MM, no {mes!r}.")
    anio, num = int(m.group(1)), int(m.group(2))

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    existentes = cargos_existentes(wb, anio, num)

    faltan, en_cero, pendientes, ok = [], [], [], []
    for tab, local in TABS.items():
        res = cargos_de_pestania(wb, tab, anio, num)
        if res is None:
            pendientes.append((local, "pestaña sin bloque de cargos"))
            continue
        calc, conteo = res
        if not calc:
            pendientes.append((local, f"sin filas {etiqueta_mes(anio, num)} en su pestaña"))
            continue
        # Regla de contrato Bigg (CONTRATOS): alquiler 50% facturado con IVA y
        # 50% "diferencia" en efectivo sin IVA, mismos montos. Si en el mes hay
        # una sola mitad, falta la de efectivo.
        if local == "Bigg" and conteo.get("Alquiler", 0) == 1:
            mitad = calc["Alquiler"][0]
            faltan.append((local, "Alquiler", mitad, 0.0,
                           "mitad en efectivo (50/50 del contrato) — faltaba"))
            print(f"⚠ Bigg {etiqueta_mes(anio, num)}: la pestaña tiene una sola mitad "
                  f"del alquiler; la mitad en efectivo (${mitad:,.2f}) también falta "
                  f"AHÍ — corregirla en la pestaña además de CARGOS.\n")
        for concepto, (monto, iva) in calc.items():
            clave = (local, concepto)
            if monto <= 0:
                continue
            if clave not in existentes:
                faltan.append((local, concepto, monto, iva,
                               "copiado de la pestaña del local"))
            elif existentes[clave][1] == 0:
                en_cero.append((local, concepto, monto, iva, existentes[clave][0]))
            else:
                dif = existentes[clave][1] - (monto + iva)
                ok.append((local, concepto, monto + iva, dif))

    print(f"CARGOS {mes} — pestañas de locales vs tabla CARGOS\n")
    if en_cero:
        print("EN CERO en CARGOS, con valor real en la pestaña del local:")
        for local, concepto, monto, iva, fila in en_cero:
            print(f"  {local:<14} {concepto:<20} ${monto:,.2f}"
                  + (f" + IVA ${iva:,.2f}" if iva else "") + f"   (fila {fila} de CARGOS)")
        print()
    if faltan:
        print("FALTAN en CARGOS:")
        for local, concepto, monto, iva, nota in faltan:
            print(f"  {local:<14} {concepto:<20} ${monto:,.2f}"
                  + (f" + IVA ${iva:,.2f}" if iva else "") + f"   ({nota})")
        print()
    if pendientes:
        print("PENDIENTES (no calculados todavía — no se inventa):")
        for local, motivo in pendientes:
            print(f"  {local:<14} {motivo}")
        print()
    if ok:
        print("YA CARGADOS (pestaña vs CARGOS):")
        for local, concepto, total, dif in ok:
            marca = "✅" if abs(dif) <= 0.01 else f"⚠ difiere ${dif:,.2f}"
            print(f"  {local:<14} {concepto:<20} ${total:>12,.2f}  {marca}")
        print()

    if not escribir:
        if faltan or en_cero:
            print("[SIN --escribir] No toqué nada. Con --escribir corrijo los EN CERO "
                  "y agrego los FALTA.")
        return

    if not (faltan or en_cero):
        print("Nada para escribir.")
        return

    sys.path.insert(0, "/Users/Facu/facu-os")
    from execution.google_auth import sheets
    s = sheets(cuenta="facu")
    fecha = f"{num}/1/{anio}"  # M/D/AAAA — Sheets en locale es-AR lo toma como 1/M
    fecha = f"1/{num}/{anio}"
    for local, concepto, monto, iva, fila in en_cero:
        s.spreadsheets().values().update(
            spreadsheetId=FILE_ID, range=f"CARGOS!D{fila}:I{fila}",
            valueInputOption="USER_ENTERED",
            body={"values": [[monto, iva, monto + iva, "", "",
                              "generar_cargos.py — corregido de 0 al valor de la pestaña"]]},
        ).execute()
        print(f"corregida fila {fila}: {local} {concepto} → ${monto + iva:,.2f}")
    if faltan:
        filas = [[fecha, local, concepto, monto, iva, monto + iva, "", "",
                  f"generar_cargos.py — {nota}"]
                 for local, concepto, monto, iva, nota in faltan]
        s.spreadsheets().values().append(
            spreadsheetId=FILE_ID, range="CARGOS!A:I",
            valueInputOption="USER_ENTERED", insertDataOption="INSERT_ROWS",
            body={"values": filas}).execute()
        print(f"agregadas {len(filas)} filas nuevas a CARGOS")


if __name__ == "__main__":
    main()

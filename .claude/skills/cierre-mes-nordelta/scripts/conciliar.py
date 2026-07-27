#!/usr/bin/env python3
"""
Conciliación del cierre de mes — Paseo Nordelta.

Los chequeos deterministas del cierre, en código y contra la fuente. Nada se
estima ni se redondea en silencio: cada diferencia se reporta al centavo.

    .venv/bin/python conciliar.py <extracto.pdf> <master.xlsx> <AAAA-MM>

Chequea:
  1. NETO del extracto (créditos − débitos) vs NETO de Movimientos (Banco, ARS).
  2. Saldo de cierre del extracto vs Banco ARS en "Saldo Actual" (solo si no hay
     movimientos de banco posteriores al mes: si los hay, lo dice y no compara).
  3. Cada movimiento del extracto (con los cargos chicos agrupados, igual que se
     cargan) presente en Movimientos, y viceversa. Match por mes + monto, como
     ya_cargados(): Facu a veces carga con otra fecha dentro del mes.
  4. Categorías huérfanas: Locales (ingresos) y Categorías (egresos) del mes que
     no existen como fila del Dashboard Mensual (ingresos 9-40, egresos 46-75).
     La comparación replica al SUMIFS real: case-insensitive y NADA más — un
     acento o un espacio de diferencia dejan la plata afuera del Dashboard, así
     que acá también cuentan como huérfana.

Sale con 0 solo si todo CIERRA. Cualquier ⚠ da código 1: es algo para mirar,
no un fallo del script. Antes de chequear valida que el parseo del PDF cierre
contra su propio saldo final y que el extracto sea del mes pedido — si no,
corta: el problema es la entrada, no la conciliación.
"""

import collections
import datetime
import pathlib
import re
import sys

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from extracto_a_movimientos import agrupar, bloque_cuenta, parsear, plata  # noqa: E402

import fitz  # noqa: E402

CENTAVO = 0.005  # "al peso" quiere decir al centavo: solo error de float


def normalizar(s):
    """Como compara el SUMIFS de verdad: solo case-insensitive.

    NO saca acentos ni espacios: en la planilla real "Categoría" y "Categoria"
    NO matchean y esa plata queda afuera del Dashboard. Si acá fuéramos más
    permisivos que el SUMIFS, el chequeo daría un falso CIERRA.
    """
    return str(s).lower()


def saldos_extracto(pdf_path):
    """(saldo último extracto, saldo final) del bloque de la cuenta en pesos."""
    doc = fitz.open(pdf_path)
    bloque = bloque_cuenta("\n".join(p.get_text() for p in doc))
    prev = final = None
    for linea in bloque.split("\n"):
        if "SALDO ULTIMO EXTRACTO" in linea:
            m = re.search(r"((?:\d{1,3}\.)*\d{1,3},\d{2})\s*$", linea)
            if m:
                prev = plata(m.group(1))
        elif "SALDO FINAL AL DIA" in linea:
            m = re.search(r"((?:\d{1,3}\.)*\d{1,3},\d{2})\s*$", linea)
            if m:
                final = plata(m.group(1))
    if prev is None or final is None:
        sys.exit("No encontré 'SALDO ULTIMO EXTRACTO' o 'SALDO FINAL AL DIA' en el PDF.")
    return prev, final


def movimientos_sheet(wb, mes):
    """Filas de Movimientos del mes: (banco_ars, todos, posteriores, descartadas).

    local/categoria se guardan CRUDOS (sin strip): el SUMIFS del Dashboard ve el
    texto tal cual está en la celda, y este script tiene que ver lo mismo.
    """
    anio, num = (int(x) for x in mes.split("-"))
    banco, todos, posteriores, descartadas = [], [], 0, 0
    acum_previo = 0.0  # neto Banco/ARS de TODOS los meses anteriores al pedido
    for r in wb["Movimientos"].iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r[:8]):
            continue
        if not isinstance(r[0], datetime.datetime):
            descartadas += 1
            continue
        clave = (r[0].year, r[0].month)
        fila = {
            "fecha": r[0].date(), "tipo": str(r[1] or "").strip().lower(),
            "medio": str(r[2] or "").strip().lower(),
            "local": str(r[3]) if r[3] is not None else "",
            "categoria": str(r[4]) if r[4] is not None else "",
            "monto": float(r[5] or 0),
            "moneda": str(r[6] or "").strip().upper(), "obs": str(r[7] or "").strip(),
        }
        es_banco_ars = fila["medio"] == "banco" and fila["moneda"] == "ARS"
        if clave == (anio, num):
            todos.append(fila)
            if es_banco_ars:
                banco.append(fila)
        elif clave > (anio, num) and es_banco_ars:
            posteriores += 1
        elif clave < (anio, num) and es_banco_ars:
            acum_previo += fila["monto"] if fila["tipo"] == "ingreso" else -fila["monto"]
    return banco, todos, posteriores, descartadas, acum_previo


def filas_dashboard(wb):
    """Etiquetas de fila del Dashboard Mensual, normalizadas → nombre exacto."""
    ws = wb["Dashboard Mensual"]
    ingresos, egresos = {}, {}
    for r in range(9, 41):
        v = ws.cell(row=r, column=1).value
        if v:
            ingresos[normalizar(v)] = str(v)
    for r in range(46, 76):
        v = ws.cell(row=r, column=1).value
        if v:
            egresos[normalizar(v)] = str(v)
    return ingresos, egresos


def multiset(movs):
    """Counter de (tipo en minúscula, monto al centavo) para cruzar presencia."""
    c = collections.Counter()
    for m in movs:
        c[(m["tipo"].lower(), round(m["monto"], 2))] += 1
    return c


def main():
    if len(sys.argv) != 4:
        sys.exit(__doc__)
    pdf, xlsx, mes = sys.argv[1], sys.argv[2], sys.argv[3]
    m = re.fullmatch(r"(\d{4})-(\d{2})", mes)
    if not m or not 1 <= int(m.group(2)) <= 12:
        sys.exit(f"El mes va como AAAA-MM (01-12), no {mes!r}.")
    anio, num = int(m.group(1)), int(m.group(2))

    crudos = parsear(pdf)
    if not crudos:
        sys.exit("El extracto parseó 0 movimientos: eso es un error, no un dato.")

    fuera_de_mes = [c for c in crudos
                    if (c["fecha"].year, c["fecha"].month) != (anio, num)]
    if fuera_de_mes:
        lineas = "\n".join(f"  {c['fecha']} {c['tipo']} ${c['monto']:,.2f} — {c['obs']}"
                           for c in fuera_de_mes)
        sys.exit(
            f"El extracto tiene {len(fuera_de_mes)} movimientos que NO son de {mes}:\n"
            f"{lineas}\n"
            f"O el mes está mal tipeado, o el PDF no es el que creés. No concilio así."
        )

    agrupados = agrupar(crudos)
    saldo_prev, saldo_final = saldos_extracto(pdf)

    # Validación del parseo: si el PDF no cierra contra sí mismo, no seguimos.
    neto_extracto = sum(c["monto"] if c["tipo"] == "Ingreso" else -c["monto"]
                        for c in crudos)
    if abs(saldo_prev + neto_extracto - saldo_final) > CENTAVO:
        sys.exit(
            f"El parseo del PDF no cierra contra su propio saldo final:\n"
            f"  {saldo_prev:,.2f} + {neto_extracto:,.2f} = "
            f"{saldo_prev + neto_extracto:,.2f} ≠ {saldo_final:,.2f}\n"
            f"Arreglar el parseo antes de conciliar nada."
        )

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    banco, todos, posteriores, descartadas, acum_previo = movimientos_sheet(wb, mes)
    if not banco:
        sys.exit(f"Movimientos no tiene ninguna fila Banco/ARS en {mes}. "
                 f"O el mes está mal, o falta cargar el extracto.")

    problemas = 0

    def check(ok, titulo, detalle_ok, detalle_mal):
        nonlocal problemas
        if ok:
            print(f"✅ {titulo}: {detalle_ok}")
        else:
            problemas += 1
            print(f"⚠ {titulo}: {detalle_mal}")

    print(f"CONCILIACIÓN {mes} — extracto {pathlib.Path(pdf).name} "
          f"({len(crudos)} movimientos, {len(agrupados)} agrupados)\n")
    print("ℹ Este script concilia SOLO la CC Especial en Pesos (la de trabajo). "
          "La CC Bancaria (VISA) del mismo PDF queda afuera: si tuvo gastos "
          "propios, hay que cargarlos aparte.\n")

    if descartadas:
        problemas += 1
        print(f"⚠ Filas descartadas: {descartadas} filas de Movimientos tienen la "
              f"fecha como texto (no como fecha) y quedaron AFUERA de todos los "
              f"chequeos. Arreglar la fecha en el sheet primero.")

    # 1. NETO
    neto_sheet = sum(f["monto"] if f["tipo"] == "ingreso" else -f["monto"]
                     for f in banco)
    dif = neto_sheet - neto_extracto
    check(abs(dif) <= CENTAVO, "NETO banco",
          f"${neto_extracto:,.2f} en los dos lados",
          f"extracto ${neto_extracto:,.2f} vs Movimientos ${neto_sheet:,.2f} "
          f"→ diferencia ${dif:,.2f}")

    # 2. Saldo de cierre
    saldo_sheet = sentinela = object()
    for row in wb["Saldo Actual"].iter_rows(min_row=1, max_row=10, values_only=True):
        if row[0] and str(row[0]).strip().lower() == "banco":
            saldo_sheet = row[1]
            break
    if saldo_sheet is sentinela:
        check(False, "Saldo de cierre", "",
              'no encontré la fila "Banco" en la pestaña Saldo Actual')
    elif saldo_sheet is None:
        check(False, "Saldo de cierre", "",
              'la fila "Banco" de Saldo Actual está VACÍA — no hay dato, '
              'que no es lo mismo que cero')
    elif posteriores:
        # Saldo Actual es el saldo de HOY: con movimientos posteriores cargados
        # no es comparable contra el cierre de este mes. Se dice, no se esquiva.
        print(f"— Saldo de cierre: no comparable — hay {posteriores} movimientos "
              f"de banco posteriores a {mes} cargados. El extracto cerró en "
              f"${saldo_final:,.2f}; verificarlo contra el extracto siguiente.")
    else:
        # Saldo Actual!B es un SUMIFS sobre TODO el historial, no el cierre del
        # mes. Si difiere, hay que saber si la diferencia es de este mes o viene
        # arrastrada de antes — se descompone en vez de culpar al mes a ciegas.
        dif = float(saldo_sheet) - saldo_final
        dif_previa = acum_previo - saldo_prev
        dif_mes = dif - dif_previa
        if abs(dif) <= CENTAVO:
            check(True, "Saldo de cierre",
                  f"${saldo_final:,.2f} en extracto y Saldo Actual", "")
        else:
            partes = [f"extracto ${saldo_final:,.2f} vs Saldo Actual "
                      f"${float(saldo_sheet):,.2f} → diferencia ${dif:,.2f}"]
            if abs(dif_previa) > CENTAVO:
                partes.append(f"${dif_previa:,.2f} vienen ARRASTRADOS de antes de "
                              f"{mes} (el acumulado del sheet difiere de la "
                              f"apertura del extracto) — buscar en meses previos")
            if abs(dif_mes) > CENTAVO:
                partes.append(f"${dif_mes:,.2f} son de {mes}")
            else:
                partes.append(f"{mes} en sí cierra al centavo")
            check(False, "Saldo de cierre", "", "; ".join(partes))

    # 3. Presencia en los dos sentidos (extracto agrupado vs sheet, por monto)
    del_extracto = multiset(agrupados)
    del_sheet = multiset(banco)
    faltan_en_sheet = del_extracto - del_sheet
    sobran_en_sheet = del_sheet - del_extracto
    detalle = []
    for (tipo, monto), n in sorted(faltan_en_sheet.items()):
        detalle.append(f"falta en Movimientos: {tipo.capitalize()} ${monto:,.2f}" +
                       (f" ×{n}" if n > 1 else ""))
    for (tipo, monto), n in sorted(sobran_en_sheet.items()):
        detalle.append(f"está en Movimientos pero no en el extracto: "
                       f"{tipo.capitalize()} ${monto:,.2f}" + (f" ×{n}" if n > 1 else ""))
    check(not detalle, "Presencia extracto ↔ Movimientos",
          f"{len(agrupados)} vs {len(banco)} movimientos, todos cruzados",
          "; ".join(detalle))

    # 4. Categorías huérfanas contra el Dashboard Mensual
    ingresos_dash, egresos_dash = filas_dashboard(wb)
    huerfanas = collections.defaultdict(float)
    for f in todos:
        if f["tipo"] == "ingreso" and f["local"].strip():
            if normalizar(f["local"]) not in ingresos_dash:
                huerfanas[("Local", f["local"])] += f["monto"]
        elif f["tipo"] == "egreso" and f["categoria"].strip():
            if normalizar(f["categoria"]) not in egresos_dash:
                huerfanas[("Categoría", f["categoria"])] += f["monto"]
    detalle = [f"{clase} {nombre!r} (${monto:,.2f}) no es fila del Dashboard"
               for (clase, nombre), monto in sorted(huerfanas.items())]
    check(not detalle, "Categorías huérfanas",
          f"los {len(todos)} movimientos del mes matchean filas del Dashboard",
          "; ".join(detalle))

    sin_etiqueta = [f for f in todos
                    if (f["tipo"] == "ingreso" and not f["local"].strip())
                    or (f["tipo"] == "egreso" and not f["categoria"].strip())]
    if sin_etiqueta:
        problemas += 1
        print(f"⚠ Sin etiqueta: {len(sin_etiqueta)} movimientos del mes sin "
              f"Local/Categoría — no suman en ningún lado del Dashboard:")
        for f in sin_etiqueta:
            print(f"    {f['fecha']} {f['tipo']} ${f['monto']:,.2f} — {f['obs']}")

    print(f"\n{'CIERRA' if not problemas else f'NO CIERRA — {problemas} punto(s) para mirar'}")
    return 1 if problemas else 0


if __name__ == "__main__":
    sys.exit(main())
